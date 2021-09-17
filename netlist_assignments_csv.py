#!/usr/bin/python3
#################################################################
#                     netlist_assignments_csv                   #
#################################################################
#                                                               #
#   This script takes an excel file that contains a loadboard   #
#   netlist within it and converts all the assignment to a csv  #  
#   in the format Net Name,Channel Assignment(s),Pin Number(s)  #
#                                                               #
#################################################################
# Version 0.0                                                   #
# By Shane Benetz                                               #
# Date: 09.15.2021                                              #
#################################################################
#################################################################
# Version 0.0 is first release 09.09.2021                       #
#################################################################

version = '0.0'

import pandas as pd
from openpyxl import load_workbook
import argparse
import os
import re
import sys
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def netlist_assignments_csv(inputFile,outputDir,productName,excluded):
    '''Searches an excel file for all the tester channel assignments for every net 
    name. The following rules must be followed in formatting the excel sheet:
        1. Only one net name/pin name per row
            - Column must be first column to have a header that has "name" in it
            - All letters must be capitalized
        2. Only one ball name/pin number per row
            - Searches for first column with 1-3 letters followed by 1-3 numbers only
            - e.g. AA12 or AB1 or A123 or P13
        3. As many channels as necessary can be added in a row
            - searches all columns in row 
            - accepted channel format examples (no names):
                a. 12345
                b. 12345-6
                c. 123-P4
                d. 123-P4-P5
                e. 123.45
                f. 231A+/231AA-
                g. PF1-PF12_NAME_345/P12_NAME_345
                h. PF1_NAME_234/P1_NAME_234
                i. any of the above proceeded by CH/TC
        note: Analog pins are defined by using the 231A+ format
        only the the letters will be translated and not the 3 numbers before 
    '''
    # check inputs/outputs
    inputFile = os.path.realpath(re.sub('["\']','',inputFile))
    if not os.path.isfile(inputFile) : 
        return print(inputFile+' is not a file')
    outputDir = os.path.realpath(re.sub('["\']','',outputDir))
    try:
        if not (os.path.isdir(outputDir)) :
            os.makedirs(outputDir)
            print('Output folder created: ',os.path.abspath(outputDir))
        outputDir = os.path.relpath(outputDir)
        if not os.access(outputDir, os.W_OK) or not os.access(outputDir, os.R_OK):
            return print('Output directory not accessible')
    except: return print('Cannot use given output directory')

    # get everything up until the first period or underscore of file name
    if productName == None: 
        productName = re.match('^(.*?)(?=(\.|_))',os.path.basename(inputFile)).group(0)
    productName = productName.replace(' ','_')

    # list out excel worksheets to select from
    try: 
        print('Loading workbook...',end='\r')
        workbook = load_workbook(filename = inputFile,data_only=True, read_only=True)
    except: return print(os.path.basename(inputFile) + ' is not a valid excel file')
    sheetsNames = workbook.sheetnames
    visibleSheets = []
    # get all the non hidden worksheets
    for sheet in sheetsNames:
        if workbook[sheet].sheet_state != 'hidden' : 
            visibleSheets.append(sheet)
    workbook.close()
    print(os.path.basename(inputFile)+'         ')
    print('Select worksheet names by typing corresponding numbers separated by spaces')
    i=0
    for sheetname in visibleSheets:
        print('    ' + str(i) + '. ' + sheetname)
        i += 1
    print('    ' + str(i) + '. ALL SHEETS')
    # get the user input
    numbers = input('Selected Numbers: ')
    numbers = re.findall(r'\d+',numbers)
    names= []; sheetNums = []
    for i in range(0,len(numbers)):
        try : 
            num = int(numbers[i].strip())
            if num == len(visibleSheets) : 
                names = visibleSheets
                for name in names:
                    if sheetsNames.index(name) not in sheetNums: 
                        sheetNums.append(sheetsNames.index(name))
                break
            if not (num >= 0 and num < len(visibleSheets)): 
                raise Exception
        except: print('Please enter only integers in above range'); return
        names.append(visibleSheets[num])
        sheetNums.append(sheetsNames.index(visibleSheets[num]))
    # pull chosen sheets from excel into raw csv files
    sheets = []
    for numb in sheetNums: 
        read_file = None
        try: # for older version on linux "sheetname" instead of "sheet_name"
            read_file = pd.read_excel(inputFile,sheetname=numb,header=None,index_col=None)
        except:
            read_file = pd.read_excel(inputFile,sheet_name=numb,header=None,index_col=None)
        fileName = sheetsNames[numb] +'_raw.csv'
        outputFileRaw = os.path.join(outputDir,fileName)
        read_file.to_csv(outputFileRaw,sep=';',mode='w')
        sheets.append(outputFileRaw)
    pNames = {}
    for sheet in sheets :
        readFile = open(sheet, 'r') 
        contents = readFile.readlines()
        nameIdxs = []
        badColNames = ['length', 'len','coord']
        badCols = []
        # try to get pin name, channel number(s), and ball number from line 
        for line in contents:
            pinName = None; channelNum = None
            pinNum = '""'
            nameIdx = None
            #change all commas in line to ! and change separators to commas
            line = line.replace(',','!').replace(';',',')
            data = line.split(',')
            if pinName == None :
                    for col in nameIdxs:
                        possName = re.search('[^a-z ]{2,}\Z',data[col])
                        if possName and len(possName.group(0))>1: 
                            pinName = possName.group(0); nameIdx = col; 
                            pinName = re.sub('[()]','',pinName); break
            for entry in data :
                updatedC = False; updatedP = False
                ind = data.index(entry)
                if ind in badCols : continue
                if any(x in entry for x in badColNames): 
                    if not ind in badCols: badCols.append(ind)
                if 'name' in entry.lower() : 
                    if not ind in nameIdxs: nameIdxs.append(ind)
                
                if pinNum == '""' and ind != nameIdx:
                    # pin number (1-3 upper case letters followed by 1-3#s) AA11
                    pin = re.match('([A-Z]{1,2}[0-9]{1,2})(?=((\Z|\s|\b)|\!))',entry)
                    if pin: 
                        pinNum = pin.group(0).strip()
                        updatedP = True
                try: 
                    # get rid of letters before channel
                    entry = re.sub('TC|CH','', entry).replace('PF','P')
                    entry = entry.replace('MCE', '231')
                    # channel formats like PF13-PF16_DPS32_425 
                    entry =  re.sub('_.*_','_', entry)
                    if '_'in entry and re.search('[0-9]{3}',entry): 
                        entry = entry[entry.find('_')+1:]+'-'+entry[:entry.find('_')]
                    # 123.45 format
                    if re.search('[0-9]{3}\.[0-9]{2}',entry) != None:
                        entry = entry.replace('.','')
                except : pass
                # 12345 or 12345-6
                channel = re.match('([0-9]{5}|([0-9]{5}-[0-9]{1,2}))(\Z|\s|\b)',entry)
                if channel: 
                    channelNum = channel.group(0).strip()
                    updatedC = True
                else:
                    #123-P4 or 123-P45 or 123-P4-P5
                    channel = re.match('[0-9]{3}-P[0-9]{1,2}(-P[0-9]{1,2})?(\Z|\s|\b)',entry)
                    if channel: 
                        channelNum = channel.group(0).strip()
                        updatedC = True
                    else:
                        #123A+ or 123HH- or CT1
                        channel = re.match('[0-9]{3}(([A-Z]{1,2}[+|-])|CT1|CT2)(\Z|\s|\b)',entry)
                        if channel: 
                            channelNum = channel.group(0).strip()
                            updatedC = True
                
                # if its a new set of assignments   
                if pinName and pinNum and channelNum and (updatedC or updatedP) and pinName!= pinNum: 
                    if pinName in pNames.keys() :
                        if channelNum in pNames[pinName] and not pinNum in pNames[pinName]:
                            pNames[pinName] += [pinNum]
                        if pinNum in pNames[pinName] and not channelNum in pNames[pinName]:
                            pNames[pinName].insert(1,channelNum)
                        elif not channelNum in pNames[pinName] and not pinNum in pNames[pinName]: 
                            pNames[pinName] += [channelNum, pinNum]
                    else: pNames[pinName] = [channelNum, pinNum]
        readFile.close()
        os.remove(sheet)
    if len(pNames.keys()) == 0 : return print('Did not find any valid assignments')
    outputFile = os.path.join(outputDir,productName+'_netlist_assignments.csv')
    writeFile = open(outputFile,'w')
    chHeader = 'Channel Number'
    testLine = pNames[list(pNames.keys())[0]]
    for i in range(0,len(testLine)):
        if re.search('[A-OQ-Z]',testLine[i]): break
        if i == 1 : chHeader = 'Channel Site-1,Channel Site-2'
        elif i > 1: chHeader += ',Channel Site-'+str(i+1)
        
    writeFile.write('Pin Name,%s,Ball Number(s)\n'%chHeader)
    for key in pNames:
        if key in excluded: continue
        writeFile.write(key+','+','.join(pNames[key])+'\n')
    writeFile.close()
    return [pNames,os.path.abspath(outputFile)]

if __name__ == '__main__' :
    parser = argparse.ArgumentParser(description=\
    '''    Searches an excel file for all tester channel assignments and every net 
    name. The following rules must be followed in formatting the excel sheet:
        1. Only one net name/pin name per row
            - Column must be first column to have a header that has "name" in it
            - All letters must be capitalized
        2. Only one ball name/pin number per row
            - Searches for first column with 1-3 letters followed by 1-3 numbers only
            - e.g. AA12 or AB1 or A123 or P13
        3. As many channels as necessary can be added in a row
            - searches all columns in row 
            - accepted channel format examples (no names):
                a. 12345
                b. 12345-6
                c. 123-P4
                d. 123-P4-P5
                e. 123.45
                f. 231A+/231AA-
                g. PF1-PF12_NAME_345/P12_NAME_345
                h. PF1_NAME_234/P1_NAME_234
                i. any of the above proceeded by CH/TC
        Note: Analog pins are defined by using the 231A+ format.
        Only the the letters will be translated and not the 3 numbers before ''', \
    formatter_class = argparse.RawTextHelpFormatter, epilog = 'usage examples:\n'\
        '   netlist_assignments_csv -i PRODUCT-NETLIST-01.xlsx -o OUTPUTDIR\n\n'\
        '   netlist_assignments_csv -i PRODUCT-NETLIST-01.xlsx -x NC N/A -n PRODUCT')
    parser.add_argument('-v', '-V', '--version', dest='version', action='store_true',\
        default=False, help='get version of script and exit')
    parser.add_argument('-i', '--input', dest='inputFile', default=None, \
        help='path of excel file to convert')
    parser.add_argument('-o', '--output', dest='outputDir', default='.', \
        help='output folder path. creates output path if DNE. DEFAULT current folder')
    parser.add_argument('-n', '--name', dest='name', default=None, \
        help='name of product and product version (e.g. fulda_B0)')
    parser.add_argument('-x', '--exclude', nargs='+',dest='exclude', default=[], \
        help='pin names to be excluded/ignored (e.g NC)')
    args = parser.parse_args()
    if args.version: print('Version '+version); sys.exit()
    try:
        netlist_assignments_csv(args.inputFile,args.outputDir,args.name,args.exclude)
    except KeyboardInterrupt:
        print('\n Keyboard Interrupt: Process Killed')
    except: print('Cannot convert given file')
