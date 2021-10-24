#!/usr/bin/python3
#################################################################
#                     stil_to_config                        	#
#################################################################
#                                                               #
#   This script takes in info from converted .stil and netlist  #
#   csv files and compares/ combines them to make a config file #
#   for the 93k advantest ATE                                   #
#                                                               #
#################################################################
# Version 0.0                                                   #
# By Shane Benetz                                               #
# Date: 09.10.2021                                              #
#################################################################
#################################################################
# Version 0.0 is first release 09.10.2021                       #
# Version 0.1 updated error logging with locations in excel     #
#################################################################

version = '0.1'

import argparse
import sys
import os
import re
import glob
import openpyxl
import itertools
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from netlist_assignments_csv import netlist_assignments_csv
from stil_assignments_csv import stil_assignments_csv

#weird but given conversions to mode value in pinconfig (value is index+1)
conversionTable = ['A+','B+','C+','D+','A-','B-','C-','D-','E+','F+','G+','H+',\
    'E-','F-','G-','H-','AA+','BB+','CC+','DD+','AA-','BB-','CC-','DD-','EE+','FF+',\
    'GG+','HH+','EE-','FF-','GG-','HH-']

# pin scale cards and their ranges of power supply channels 
PS9G = ['31701-32416', '30101-30216', '30501-30616']
DCS_DPS128HC = ['22501-22516','42501-42516','22901-22916','42901-42916']
PS1600 = ['12501-13216', '11701-12416', '10101-10816', '20101-20816', '40101-40816',\
    '10901-11616', '20901-21616', '40901-41616', '21701-22416', '41701-42416']
DCS_UHC4T = ['22701-22704', '23001-23004']
thinBorders = Border(left=Side(style='thin'),right=Side(style='thin'), \
    top=Side(style='thin'), bottom=Side(style='thin'))

def stil2config(inputFiles, outputDir, productName, card, anType, printErr):
    '''    Takes in data containing pin defintion for a device and creates a .conf file
    with all the provided information. Can accept the CSVs created by 
    "still_assignment_csv" and/or "netlist_assignments_csv" or it can accept
    the actual excel netlist and/or .stil file(s) themselves'''
    stilFiles = []; netlistFile = None; netlistCSV = None; stilCSV = None
    fileTypes = ['.xslx','.xls','.xlsm','.stil','assignments.csv']
    finalFiles = []
    inDir = None
    for file in inputFiles:
        if os.path.isdir(file): 
            inDir = file
            inputFiles.extend(glob.glob(file+ '/*'))
        if any(file.find(x) != -1 for x in fileTypes) :
            finalFiles.append(file)
    if len(finalFiles) == 0 and inDir: 
        s = 'No valid assignment files in %s or listed files'%os.path.abspath(inDir)
        return print(s)
    for file in finalFiles:
        file = os.path.realpath(re.sub('["\']','',file))
        if not os.path.isfile(file):
            print(file+' is not a file'); continue
        if file.endswith('.stil'): stilFiles.append(file)
        elif file.endswith('stil_assignments.csv'): stilCSV = file
        elif file.endswith('netlist_assignments.csv'): netlistCSV = file
        elif re.search('.xslx|.xls|.xlsm',file): netlistFile = file
    outputDir = os.path.realpath(re.sub('["\']','',outputDir))
    try:
        if not (os.path.isdir(outputDir)) :
            os.makedirs(outputDir)
            print('Output folder created: ',os.path.abspath(outputDir))
        outputDir = os.path.relpath(outputDir)
        if not os.access(outputDir, os.W_OK) or not os.access(outputDir, os.R_OK):
            return print('Output directory not accessible')
    except: return print('Cannot use given output directory')

    if productName == None: # everything until the first period or underscore
        fileName = netlistCSV if netlistCSV else netlistFile 
        if fileName == None:
            return print('No netlist assignments found. Please provide either '\
                'netlist CSV or netlist excel file')
        productName = re.match('^(.*?)(?=(\.|_))',os.path.basename(fileName)).group(0)
        productName = productName.replace(' ','_')
    # get netlist assignments
    netDict = None
    ballMap = None
    locations = {}
    if netlistCSV == None and netlistFile :
        netDict, netlistCSV, ballMap = netlist_assignments_csv(netlistFile,outputDir,productName,[None])
    if netlistCSV :
        with open(netlistCSV,'r') as netlist:
            lines = netlist.readlines()
            if not 'Pin Name,Channel' in lines[0]:
                return print('\nCannot find valid netlist assignments')
            tempDict = {}
            for line in lines[1:]:
                if len(line) < 4 or line.count(',')<2: break
                try:
                    data = line.replace('\n','').split(',')
                    tempDict[data[0]] = data[1:]
                    if ':Row' in data[-1] : 
                        locations[data[0]] = data[-1]
                        if data[-2].startswith('(') and data[-2].endswith(')'):
                            locations[data[0]] = data[-2]+','+locations[data[0]]
                        data.pop(-1)
                except: pass
            netDict = tempDict
    if netDict == None or netlistCSV == None : 
        return print('\nCannot find valid netlist assignments')
    
    # get .stil assignments
    stilList = None
    if stilCSV == None and len(stilFiles)>0:
        stilList, stilCSV = stil_assignments_csv(stilFiles,outputDir,productName)
    if stilCSV and stilList == None:
        with open(stilCSV,'r') as stil:
            lines = stil.readlines()
            if not 'Pin Name,In/Out/InOut' in lines[0]:
                stilCSV == None
                print('\nCannot find valid .stil assignments')
            tempList = []
            for line in lines[1:]:
                if len(line) < 4: break
                tempList.append(line[:line.find('\n')])
            stilList = tempList
    if stilList == None or stilCSV == None : 
        print('\nCannot find valid .stil files, getting all assignments from netlist'\
            ' names (All pins IO)')
        tempFile = os.path.join(outputDir,'temp_stil_file.stil')
        with open(tempFile,'w') as tempStil:
            tempStil.write('Signals {\n')
            for key in netDict.keys():
                tempStil.write(key+' InOut; \n')
            tempStil.write('}')
        stilList, stilCSV = stil_assignments_csv([tempFile],outputDir,productName)
        os.remove(tempFile)

    #get number of sites from the netlist CSV header
    with open(netlistCSV,'r') as net:
        topLine = net.readline()
        numbers = re.findall(r'\d+',topLine) # get biggest integer on top line
        if len(numbers) == 0: sites = 1
        else: sites = int(max(numbers))

    diffs, stilDiffs, netDiffs = get_diff(netDict,stilList,locations)
    # check for channels that dont have ball assignment: trigger channels
    extraCONFI = []
    try:
        #extraInNet = diffs[diffs.index('\nIn netlist but not in .stil:')+1:]  
        for net in netDiffs:
            if '""' in netDict[net] : 
                stilList.append(net+',I'); extraCONFI.append(net)
                diffs.remove(net)
    except:pass  

    #find if there are any names that have only be changed a bit netween stil & net
    transferNames = {}
    for sName in stilDiffs:
        if sName in netDict.keys(): continue
        for nName in netDiffs:
            #reverse the string then look for the first group of numbers, then 
            #reverse the string back and convert to an int
            try:
                stilNums = None; netNums = None
                stilNums = re.search('\d+',sName[::-1])
                if stilNums: stilNums = int(stilNums.group(0)[::-1])
                else: stilNums = ''
                netNums = re.search('\d+',nName[::-1])
                if netNums: netNums = int(netNums.group(0)[::-1])
                else: netNums = ''
                if stilNums != netNums: continue
            except: pass
            #get just the letters
            stilChars = None ; netChars = None
            stilChars = re.sub('[\[\]]','',sName).split('_')
            netChars = re.sub('[\[\]]','',nName).split('_')
            testList = [] 
            #check each chunk of var name to see if stil is a variant of net name
            for Schunk in stilChars:
                for Nchunk in netChars:
                    if Nchunk in testList: continue
                    #if the netlist chnk is substring of stil list chunk
                    if Schunk == Nchunk or(len(Nchunk) > 1 and Nchunk in Schunk): 
                        if not Nchunk in testList: testList.append(Nchunk); break
            if testList == netChars and testList != []:
                transferNames[nName] = [sName]
    #start writing to .conf file
    configFileName = os.path.relpath(os.path.join(outputDir,productName+'.conf'))
    configFile = open(configFileName,'w')
    configFile.write('hp93000,config,0.1\n')
    entries = []
    names = list(netDict.keys())
    names += stilList
    PSs = []
    pinCounts = {}
    for pinName in names:
        stilName = None
        if pinName.find(',')>0: # name has IO info in it
            stilName = pinName
            pinName = pinName[:pinName.find(',')]
        if not pinName in netDict.keys(): continue #not in netlist
        if pinName in pinCounts.keys(): continue
        pinData = netDict[pinName]
        #skip names that dont have coherent assignments
        skip = False; matched = False
        for datum in pinData: # case of channel, ball, channel, ball, etc
            if re.match('([A-Z]{1,2}[0-9]{1,2})|(\"\")',datum): matched = True
            if matched and re.match('^\d{3}',datum): skip = True
        if skip : continue
        pins = [x for x in pinData if re.match('([A-Z]{1,2}[0-9]{1,2})|(\"\")',x)] 
        channels = [y for y in pinData if re.match('^\d{3}',y)] 
        analog = False
        #convert the channels to 5 number format
        for i in range(0,len(channels)):
            channel = channels[i]
            # for analog pins that need the conversion (231A+ etc.)
            # analog pins can only have one definition
            if re.match('[0-9]{3}(([A-Z]{1,2}[+|-])|CT1|CT2)(\Z|\s|\b)',channel):
                analog=True
                for pad in conversionTable: #convert letters to pin #
                    if re.sub('[0-9]','',channel).strip() != pad: continue
                    if len(pad) == 3: io='o'
                    else:io='i'
                    if len(pins) == 1 and re.match('([A-Z]{1,2}[0-9]{1,2})',pins[0]):
                        ball = pins[0]
                    else: ball = ''
                    if stilName:
                        outString = 'DFAN "%s231,%d,%s","%s",(%s)'%    \
                            (anType,conversionTable.index(pad)+1,io,ball,pinName)
                        entries.append(outString); break
            # format 123-P4 --> [12304]    
            elif re.match('[0-9]{3}-P[0-9]{1,2}(\Z|\s|\b)',channel):
                channelTemp = channel[0:3] + '%02d' % int(channel[5:])
                channels[i] = [channelTemp]
            # format 123-P4-P5 --> [12304,12305] 
            elif re.match('[0-9]{3}-P[0-9]{1,2}-P[0-9]{1,2}(\Z|\s|\b)',channel) :
                start = int(channel[channel.find('P')+1:channel.rfind('-')])
                end = int(channel[channel.rfind('P')+1:])
                channelTemp = []
                for j in range(start, end+1):
                    channelTemp.append(channel[0:3]+'%02d' % j)
                channels[i] = channelTemp
            # format 12345-6 --> [12345,12346]
            elif re.match('[0-9]{5}-[0-9]{1,2}(\Z|\s|\b)',channel) :
                start = int(channel[3:5])
                end = int(channel[channel.find('-')+1:])
                channelTemp = []
                for j in range(start, end+1):
                    channelTemp.append(channel[0:3]+'%02d' % j)
                channels[i] = channelTemp
            # already in format [12345]
            else: channels[i] = [channel]
        if analog: continue
        i=0
        isPS = False
        if pinName in transferNames.keys() and channels != None:
            for ch in channels:
                transferNames[pinName] += ch
        for chs in channels:
            i+=1
            for ch in chs:
                if not(ch.isdigit() and len(ch) == 5): continue
            if len(chs) ==1: chnlString = chs[0]
            else: chnlString = '(%s)'%','.join(chs)
            if i == 1: # first definition of a channel
                # there are more than one ball numbers associated = power supply
                if len(pins)>1 : 
                    entry = 'DFPS ' + chnlString +',POS,('+pinName +')'
                    if not entry in entries:
                        entries.append(entry); isPS = True; pinCounts[pinName] = 1
                # there are multiple channels in assignment [12345,12346]
                elif len(chs) > 1: 
                    entry = 'DFPS ' + chnlString +',POS,('+pinName +')'
                    if not entry in entries:
                        entries.append(entry); isPS = True; pinCounts[pinName] = 1
                else:
                    lists = [DCS_DPS128HC,DCS_UHC4T]
                    if 'PS9G' in card: lists.append(PS9G)
                    if 'PS1600' in card: lists.append(PS1600)
                    # check to see if the channel is within the power supply ranges
                    for l in lists:
                        for ranges in l: # from range lists at top of doc
                            low = int(ranges[:ranges.find('-')])
                            high = int(ranges[ranges.find('-')+1:])
                            firstCh = int(chs[0])
                            if firstCh <= high and firstCh >= low : #if in PS range
                                entry = 'DFPS ' + chnlString +',POS,('+pinName +')'
                                if not entry in entries:
                                    entries.append(entry); pinCounts[pinName] = 1
                                isPS = True; break
                # not a power supply, just a regular pin 
                if not isPS and ((stilName and stilName in stilList) \
                    or pinName in transferNames.keys()): 
                    entry = 'DFPN %s,"%s",(%s)'%(chnlString,pins[0],pinName)
                    #if its a trigger
                    entry = entry.replace('""""','""')
                    if not entry in entries:
                        entries.append(entry); pinCounts[pinName] = 1
                elif not isPS: break
            # define PALS for multi-site
            if i>1 :  
                entry = 'PALS %d,%s,,(%s)'%(i,chnlString,pinName)
                if not entry in entries and pinName in pinCounts.keys():
                    entries.append(entry); pinCounts[pinName] += 1
        if isPS and pinName not in PSs: PSs.append(pinName) # track power supplies
    if len(PSs) > 0: entries.append('CONF DC,POWER,(%s)'%','.join(PSs))
    entries.append('PSTE '+str(sites))
    #write the transfer file
    transferFile =  os.path.join(outputDir,productName+'_transfer_names.txt')
    if len(transferNames.keys()) > 0:
        with open(transferFile,'w') as tf:
            tf.write('Stil Name --> Netlist Name, Channel\n')
            for name in transferNames.keys():
                tnList = transferNames[name]
                stilName = tnList[0]
                tnList[0] = name
                tf.write(stilName + ' --> ' + ', '.join(tnList)+ '\n')
                tnList[0] = stilName
    # get all the group/conf definitions from the .stil conversion csv
    with open(stilCSV,'r') as groupsFile:
        content = groupsFile.read()
        groups = content[content.find('#'):]
        for nName in transferNames.keys():
            sName = transferNames[nName][0]
            try:
                diffInd = diffs.index(sName)
                diffs[diffInd] = diffs[diffInd]+' (transfer found)'
            except: pass
            groups = groups.replace(','+sName+',',','+nName+',')
            groups = groups.replace(','+sName+')',','+nName+')')
            groups = groups.replace('('+sName+',','('+nName+',')
        for stilName in diffs: # replace ones that dont exist in stillist
            groups = groups.replace(','+stilName+',',',')
            groups = groups.replace(','+stilName+')',')')
            groups = groups.replace('('+stilName+',','(')
    if len(extraCONFI)>1: entries.append('DFGP I,(%s)(triggers)'%','.join(extraCONFI))
    groups = groups.splitlines()
    confNames = [x[x.rfind('(')+1:x.rfind(')')] for x in entries if x.startswith('DFPN')]
    #rename groups if necessary
    for group in groups:
        gInd = groups.index(group)
        gNames = group[group.find('(')+1:group.find(')')].split(',')
        gNames = [x for x in gNames if x in confNames]
        if group.startswith('CONF'): 
            groups[gInd] = groups[gInd][:groups[gInd].find('(')+1] + ','.join(gNames) + ')'
            continue
        largestSub = (os.path.commonprefix(gNames)).lower()
        if largestSub.endswith('_') or largestSub.endswith('['): 
                largestSub = largestSub[0:len(largestSub)-1]
        if len(largestSub) < 3 or len(gNames) < 2: groups[gInd] = None
        else:
            groups[gInd] = groups[gInd][:groups[gInd].rfind('(')+1] + largestSub + ')' 
    for group in groups:
        if group == None: continue
        if group.startswith('CONF I,') and len(extraCONFI) > 0:
            group = group[:group.rfind(')')]+','+','.join(extraCONFI)+')'
        # if (group.startswith('CONF') or group.startswith('DFGP')) and \
        #     group[group.find('('):group.find(')')].count(',')>1 :
        entries.append(group)      
    order = ['DFPN','DFPS','DFAN','PALS','PSTE','CONF','DFGP']
    entries = sorted(entries,key=lambda x:(order.index(x[0:4]),x[x.rfind('('):]))

    #print differences to error log
    errlogFile = os.path.join(outputDir,productName+'_config_error_log.txt')
    if os.path.isfile(errlogFile): os.remove(errlogFile)
    if len(diffs) > 2: 
        with open(errlogFile,'w') as err:
            err.write('\n'.join(diffs).strip())

    # log repeated elements from config  
    definitions = entries[0:entries.index('PSTE '+str(sites))]
    oddities = find_oddities(definitions,pinCounts,sites,locations)
    configFile.write('\n'.join(entries))
    configFile.write('\nNOOP "7.4.2",,,')
    configFile.close()
    if len(oddities) > 0:
        with open(errlogFile,'a') as err:
            err.write('\n' + '\n'.join(oddities).strip())
    #print error log to terminal
    if os.path.isfile(errlogFile):
        if printErr:
            with open(errlogFile,'r') as err:
                print(err.read().strip())
        print('Error log location: ',os.path.relpath(errlogFile))
    else:
        print('\nNo Issues found')
    if os.path.isfile(transferFile):
        if printErr:
            with open(transferFile,'r') as tran:
                print(tran.read().strip())
        print('Pin Name Cross-Refs location: ',os.path.relpath(transferFile))
    print('Config file location: '+ '\x1b[0;30;43m' +\
            configFileName + '\x1b[0m')
    #print(ballMap)
    make_excel_docs(productName,outputDir,netDict,configFileName,ballMap)

def find_oddities(entries,pinCounts,sites,locations):
    words = {}; oddities = []
    for line in entries:
        name = line[line.rfind('(')+1:line.rfind(')')]
        if name in locations.keys():
            if line.startswith('PALS') and ',' in locations[name]:
                loc = locations[name][locations[name].find(',')+1:]
            elif line.startswith('DF') and ',' in locations[name]:
                loc = locations[name][:locations[name].find(',')]
            else:
                loc = locations[name]
        else : loc = ''
        #get the channel assignments and the ball assignments
        defs = re.findall('.\d{5},',line)+re.findall('".*?"',line)
        for nam in defs:
            item = (nam[1:-1] if len(nam)>2 else None)
            if item == None or (item and len(item) < 2) or item == '': continue
            if item in words.keys(): 
                if not '\n\nRepeated Definitions:' in oddities:
                    oddities.append('\n\nRepeated Definitions:')
                oddities.append('Repeated "%s" on .conf line %d %s, first occurance .conf line %s'%\
                    (item,entries.index(line)+2,loc,words[item]))
            else: words[item] = str(entries.index(line)+2) + ' ' + loc
    
    for pin in pinCounts.keys():
        if pinCounts[pin] != sites:
            if not '\n\nUnusual Occurances:' in oddities:
                oddities.append('\n\nUnusual Occurances:')
            errString = '%d occurence(s) of %s when there should be %d'% \
                (pinCounts[pin], pin, sites)
            oddities.append(errString)
    return oddities

def get_diff(netDict,stilList,locations):
    differ = ['In .stil but not in netlist:']
    stilListNames = []
    stilDiff = []; netDiff = []
    for assignment in stilList :
        stilListNames.append(assignment[:assignment.find(',')])
    netDictNames = netDict.keys()
    for name in stilListNames:
        if name not in netDictNames: differ.append(name); stilDiff.append(name)
    differ.append('\nIn netlist but not in .stil:')
    for name in netDictNames:
        if name in locations.keys(): loc = locations[name]
        else : loc = ''
        if name not in stilListNames: differ.append(name+' '+loc); netDiff.append(name)
    return differ, stilDiff, netDiff

def make_excel_docs(productName, outputDir, netDict, configFile, ballMap):
    outFileName = os.path.join(outputDir,productName+'config_info.xlsx')
    i = 0
    while os.path.isfile(outFileName) :
        i+=1
        outFileName = outFileName[:outFileName.rfind('_info')] + '_info_' +str(i) + '.xlsx'
    
    workbook = openpyxl.Workbook()
    
    for worksheet in workbook.sheetnames : 
        if workbook[worksheet].max_column == 1 and workbook[worksheet].max_row == 1:
            del workbook[worksheet] # delete extra empty sheets
    WS = workbook.create_sheet('Ball Map')
    assignments = {}
    if ballMap:
        for ballLoc in ballMap.keys():
            if ballLoc == '""':continue
            try:
                WS[ballLoc] = ballMap[ballLoc]
            except:pass
    else:
        with open(configFile,'r') as config:
            content = config.read()
        for pinName in netDict.keys(): 
            if content.find('('+pinName+')') < 0 : continue
            line = ','.join(netDict[pinName]) + ','
            balls = re.findall(',[A-Z]{1,3}[0-9]{1,3},',line)
            channel = line[:line.find(',')]
            for ball in balls:
                assignments[ball.replace(',','')] = pinName  + '                 ' + channel
        for ballLoc in assignments.keys():
            if ballLoc == '""':continue
            WS[ballLoc] = assignments[ballLoc] 
    #set dimensions
    for col in range(WS.max_column+1):
        WS.column_dimensions[get_column_letter(col+1)].width = 13
    for row in range(WS.max_row+1):
        WS.row_dimensions[row].height = 64   
    #add borders
    for row in WS.rows:
        for cell in row:
            cell.border = thinBorders
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    #delete empty columns
    for col in range(1,WS.max_column+1):
        empty = True
        for row in range(1,WS.max_row+1):
            if WS.cell(row=row,column=col).value: empty = False; break
        if empty:
            WS.delete_cols(col,1)
    workbook.save(outFileName)


    
if __name__ == '__main__' :
    parser = argparse.ArgumentParser(description=\
    '''    Takes in data containing pin defintion for a device and creates a .conf 
    file with all the provided information. Can accept the CSVs created by 
    "still_assignment_csv" and/or "netlist_assignments_csv" or it can accept
    the actual excel netlist and/or .stil file(s) themselves''', \
    formatter_class = argparse.RawTextHelpFormatter, epilog = 'usage examples:\n'\
        '   stil2config -i netlist.xlsx product_stil_assignments.stil -o productDir\n\n'\
        '   stil2config -i netlist_assignments.csv -o productDir -n product -p\n\n'\
        '   stil2config --io folder_with_xslx&stil -n product -c PS1600 -a MCB\n\n')
    parser.add_argument('-v', '-V', '--version', dest='version', action='store_true',\
        default=False, help='get version of script and exit')
    parser.add_argument('-i', '--input', nargs='+', dest='inputs', default=['.'], \
        help='path of excel file and .stil file(s) to convert or the already converted'\
            '\ncsv version of both. can also take directory with files in it')
    parser.add_argument('-o', '--output', dest='outputDir', default='.', \
        help='output folder path. creates output path if DNE. DEFAULT current folder')
    parser.add_argument('--io', dest='inOut', default=None,\
        help='path to input and output directory if they are the same')
    parser.add_argument('-n', '--name', dest='name', default=None, \
        help='name of product and product version (e.g. fulda_B0)')
    parser.add_argument('-c', '--cards', dest='psCard', default=['PS9G'],\
        choices=['PS9G', 'PS1600'], nargs='+',
        help='which pin scale card(s) are being used. DEFAULT: PS9G only')
    parser.add_argument('-a', '--analog', dest='ana', default='MCE',\
        choices=['MCE', 'MCB','MCA'],
        help='which analog card is being used. DEFAULT: MCE')
    parser.add_argument('-p', '--print', dest='printerr', default=False,\
        action='store_true',help='print error log to terminal')
    args = parser.parse_args()
    if args.version: print('Version '+version); sys.exit()
    if args.inOut != None :
        if not os.path.isdir(args.inOut) : 
            print(args.inOut,' is not a directory'); sys.exit()
        args.inputs = [args.inOut]
        args.outputDir = args.inOut
    try:
        stil2config(args.inputs,args.outputDir,args.name,args.psCard,args.ana,\
            args.printerr)
    except KeyboardInterrupt:
        print('\n Keyboard Interrupt: Process Killed')
    #except: print('Cannot convert given files')